import sqlite3
import pandas as pd
import smtplib

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook

### ---- CONFIG ----
DB_PATH = "data.db"
TABLE_NAME = "REL"
EMAILS_FILE = "emails.txt"

SMTP_HOST = "smtp.vitesco-technologies.net"
SMTP_PORT = 587
SMTP_USER = "svv33684@vitesco.com"
SMTP_PASS = "Schaeffler.2025"
FROM_EMAIL = "calibration.maintenance@vitesco-technologies.net"
### ---- END CONFIG ----

def get_emails():
    with open(EMAILS_FILE, encoding="utf8") as f:
        return [l.strip() for l in f if "@" in l]

def send_email(to_addr, subject, message, html_part=None):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = FROM_EMAIL
    msg["To"] = to_addr
    part1 = MIMEText(message, "plain", "utf-8")
    msg.attach(part1)
    if html_part:
        part2 = MIMEText(html_part, "html", "utf-8")
        msg.attach(part2)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(FROM_EMAIL, [to_addr], msg.as_string())
        print(f"Sent to {to_addr}: {subject}")
    except Exception as e:
        print(f"EROARE la trimitere email către {to_addr}: {e}")

def get_cell_bg_colors(excel_path, sheet_name, match_columns):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    result = {}  # (row_idx, colname) -> color_hex
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for colidx, cell in zip(range(len(headers)), row):
            colname = headers[colidx]
            if colname in match_columns:
                fill = cell.fill
                color = None
                if fill and fill.fgColor:
                    rgb = fill.fgColor.rgb
                    if rgb and isinstance(rgb, str) and len(rgb) == 8:
                        color = '#' + rgb[-6:]
                result[(cell.row, colname)] = color
    return result

def main():
    today = pd.Timestamp.today().normalize()
    recipients = get_emails()

    df = pd.read_sql(f"SELECT * FROM {TABLE_NAME}", sqlite3.connect(DB_PATH))
    notif_1d, notif_7d, notif_1m = [], [], []

    # Citește culorile reale din Excel
    EXCEL_PATH = DB_PATH.replace(".db", ".xlsx")
    SHEET_NAME = TABLE_NAME
    KEY_COLUMNS = ["Calibration Due date", "Maintenance planning"]
    colors = get_cell_bg_colors(EXCEL_PATH, SHEET_NAME, KEY_COLUMNS)

    for idx, row in df.iterrows():
        equip = row.get("Equipment Identification", "Unknown equipment")
        for field, label in [
            ("Calibration Due date", "Calibration date"),
            ("Maintenance planning", "Maintenance date"),
        ]:
            raw_val = row.get(field)
            if not raw_val or pd.isna(raw_val):
                continue
            try:
                dt = pd.to_datetime(raw_val, errors="coerce")
                if pd.isnull(dt): continue
                zile_ramase = (dt.normalize() - today).days

                # Culoare din Excel
                row_excel = idx + 2  # header = 1, idx e zero-based
                color_hex = colors.get((row_excel, field), None)
                symbol_unicode = "⚪"
                symbol_html = '<span style="color:#bbb;font-size:1.45em;vertical-align:middle;">●</span>'  # gri

                if color_hex:
                    if color_hex.lower() == "#ff0000":
                        symbol_unicode = "🔴"
                        symbol_html = '<span style="color:#f00;font-size:1.45em;vertical-align:middle;">●</span>'
                    elif color_hex.lower() == "#00ff00":
                        symbol_unicode = "🟢"
                        symbol_html = '<span style="color:#2cb526;font-size:1.45em;vertical-align:middle;">●</span>'

                msg = f"{symbol_unicode} {equip}: {label}: {dt.date()}"

                html_msg = f"{symbol_html} <b>{equip}</b>: {label}: <b>{dt.date()}</b>"

                if zile_ramase in [30, 31]:
                    notif_1m.append((msg, html_msg))
                elif zile_ramase == 7:
                    notif_7d.append((msg, html_msg))
                elif zile_ramase == 1:
                    notif_1d.append((msg, html_msg))
            except Exception as ex:
                print(f"ERROR processing row {idx} [{field}]: {ex}")

    notifications_exist = notif_1d or notif_7d or notif_1m
    if notifications_exist:
        subject = "Equipment Calibration / Maintenance Reminder"

        # PLAIN TEXT
        message = "Hello,\n\nThe following equipment requires attention:\n\n"
        if notif_1d:
            message += "!!! TOMORROW !!!\n"
            for msg, _ in notif_1d:
                message += "- " + msg + " (1 day left)\n"
            message += "\n"
        if notif_7d:
            message += "!! In 7 days:\n"
            for msg, _ in notif_7d:
                message += "- " + msg + " (7 days left)\n"
            message += "\n"
        if notif_1m:
            message += "In one month:\n"
            for msg, _ in notif_1m:
                message += "- " + msg + " (1 month left)\n"
            message += "\n"
        message += "This message was sent automatically by the equipment management system."

        # HTML cu buline colorate!
        html = """
        <div style="font-family:Arial,sans-serif;">
        <p>Hello,</p>
        <p>The following equipment requires attention:</p>
        """

        if notif_1d:
            html += '<div style="margin-bottom:13px;">'
            html += '<div style="font-weight:bold; font-size:1.12em; margin-bottom:4px; border-bottom:2px solid #ddd; padding-bottom:4px;">!!! TOMORROW !!!</div>'
            html += '<ul style="margin-top:2px; margin-bottom:5px;">'
            for _, html_msg in notif_1d:
                html += f'<li style="font-weight:bold;">{html_msg} <span>(1 day left)</span></li>'
            html += '</ul></div>'

        if notif_7d:
            html += '<div style="margin-bottom:13px;">'
            html += '<div style="font-weight:bold; font-size:1.10em; border-bottom:2px dotted #ccc; margin-bottom:4px; padding-bottom:2px;">!! In 7 days</div>'
            html += '<ul style="margin-top:2px; margin-bottom:5px;">'
            for _, html_msg in notif_7d:
                html += f'<li style="font-weight:bold;">{html_msg} <span>(7 days left)</span></li>'
            html += '</ul></div>'

        if notif_1m:
            html += '<div style="margin-bottom:13px;">'
            html += '<div style="font-weight:bold; font-size:1.06em; border-bottom:1px solid #eee; margin-bottom:4px;">In one month</div>'
            html += '<ul style="margin-top:2px; margin-bottom:5px;">'
            for _, html_msg in notif_1m:
                html += f'<li style="font-weight:bold;">{html_msg} <span>(1 month left)</span></li>'
            html += '</ul></div>'

        html += """<p style="margin-top:14px; color:#555;">This message was sent automatically by the equipment management system.</p></div>"""

        for email in recipients:
            send_email(email, subject, message, html_part=html)
            print(f"Sent to {email}: {subject}")
    else:
        print("No reminders for today.")

if __name__ == "__main__":
    main()