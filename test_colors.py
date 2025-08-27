# test_colors.py
EXCEL_PATH = "data.xlsx"
SHEET_NAME = "REL"

def test_excel_colors():
    from openpyxl import load_workbook
    import os

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    print(f"Tabel de culori din: {os.path.basename(EXCEL_PATH)}")

    for i, row in enumerate(ws.iter_rows()):
        color_line = []
        for cell in row:
            fill = cell.fill
            col = None
            if fill and fill.fgColor:
                fg = fill.fgColor
                if fg.type == "rgb" and fg.rgb:
                    col = '#' + fg.rgb[-6:]
                elif fg.type == "theme":
                    col = "#DDEEFF"
            color_line.append(col if col else "—")
        print(f"Rând {i + 1}: {color_line}")

# Rulează testul
test_excel_colors()
