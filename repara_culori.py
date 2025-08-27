from openpyxl import load_workbook
import pandas as pd
import sqlite3

EXCEL_PATH = 'data.xlsx'
DB_PATH = 'data.db'
SHEET_NAME = 'REL'

wb = load_workbook(EXCEL_PATH)
ws = wb[SHEET_NAME]
headers = [str(cell.value).strip() for cell in ws[1]]

conn = sqlite3.connect(DB_PATH)
df = pd.read_sql(f"SELECT * FROM {SHEET_NAME}", conn)
for col_idx, col_name in enumerate(headers):
    color_col = f"{col_name}_color"
    if color_col not in df.columns:
        df[color_col] = None
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx + 1)
        fill = cell.fill
        color_string = None
        if fill and fill.fill_type == "solid":
            fg = fill.fgColor
            if hasattr(fg, "rgb") and fg.rgb:
                color_string = "#" + fg.rgb[-6:]
            # indexed? dacă ai nevoie, adaugă aici.
        if pd.isna(df.at[row - 2, color_col]) or not df.at[row - 2, color_col]:
            df.at[row - 2, color_col] = color_string
df.to_sql(SHEET_NAME, conn, if_exists='replace', index=False)
conn.close()
print("S-au reimplantat culorile inițiale acolo unde s-au pierdut/erau None.")