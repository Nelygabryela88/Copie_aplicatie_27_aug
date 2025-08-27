import os
import pandas as pd
import sqlite3
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from werkzeug.utils import secure_filename
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
import threading
from flask import request, jsonify
import os
from openpyxl.comments import Comment
excel_write_lock = threading.Lock()
app = Flask(__name__)
app.secret_key = "Schaeffler"

EXCEL_PATH = 'data.xlsx'
DB_PATH = 'data.db'
SHEET_NAME = 'REL'
HISTORY_PATH = 'history.log'
def get_active_excel_path():
    file_id = session.get('file_id', None)
    file_name = session.get('file_name', None)
    if file_id and file_name:
        filename = f"files/{file_id}_{file_name}"
        if os.path.exists(filename):
            return filename
    return EXCEL_PATH  # fallback la fișierul default
def log_change(action, rowid, column=None, value=None):
    try:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(HISTORY_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {action} | rowid={rowid} | column={column} | value={value}\n")
    except Exception as e:
        print(f"Error logging change: {e}")
def get_active_sheet():
    return session.get('sheet', SHEET_NAME)
#def init_db():
    #df = read_excel_file()
    #if df is not None:
        #print("Dataframe loaded successfully")
        #with sqlite3.connect(DB_PATH) as conn:
            #conn.execute(f'DROP TABLE IF EXISTS {SHEET_NAME}')
            #df.to_sql(SHEET_NAME, conn, if_exists='replace', index=False)
           # print("Dataframe stored in database")
    #else:
        #print("Failed to load dataframe") 
def init_db():
    # Încarcă direct EXCELUL DEFAULT
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, header=0)
        df.dropna(how='all', inplace=True)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

        print("Dataframe loaded successfully")
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(f'DROP TABLE IF EXISTS {SHEET_NAME}')
            df.to_sql(SHEET_NAME, conn, if_exists='replace', index=True, index_label='rowid')
            print("Dataframe stored in database")
    except Exception as e:
        print("Failed to load dataframe:", e)

def read_excel_file():
    excel_path = get_active_excel_path()
    sheet_name = get_active_sheet()
    # 1. Citești date ca DataFrame
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=0)
    df.dropna(how='all', inplace=True)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    
    # 2. Citești workbook-ul cu openpyxl
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    headers = [str(cell.value).strip() for cell in ws[1]]

    # 3. Pentru fiecare coloană, pentru fiecare celulă, ia "fill" și păstrează-l în <col>_color
    for col_idx, col_name in enumerate(headers):
     color_values = []
     for row in range(2, ws.max_row+1):
        cell = ws.cell(row=row, column=col_idx+1)
        fill = cell.fill
        color_string = None
        if fill is not None and fill.fill_type == "solid":
            fg = fill.fgColor
            if fg is not None:
                # Cazul clasic: RGB
                if hasattr(fg, "rgb") and fg.rgb:
                    color_string = "#" + fg.rgb[-6:]
                # Cazul INDEXED
                elif hasattr(fg, "indexed") and fg.indexed is not None:
                    # Mapare minimală, adaugă/ajustează după nevoie
                    indexed2rgb = {
                        2: "#ffffff",   # White
                        3: "#000000",   # Black
                        10: "#ff0000",  # Red
                        22: "#00ff00",  # Green
                        41: "#0000ff",  # Blue
                        # ...poți extinde lista, dacă ai alte nuanțe...
                    }
                    color_string = indexed2rgb.get(fg.indexed, None)
                # Caz theme - ignorăm deocamdată
                elif hasattr(fg, "theme") and fg.theme is not None:
                    color_string = None
        color_values.append(color_string)
        df[col_name + "_color"] = color_values
    return df

def get_excel_colors(columns_filter):
    try:
        excel_path = get_active_excel_path()
        sheet_name = get_active_sheet()
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]

        excel_headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        idx_map = [excel_headers.index(col) if col in excel_headers else -1 for col in columns_filter]

        colors = []
        for r in range(2, ws.max_row + 1):
            row_colors = []
            for i in idx_map:
                if i == -1:
                    row_colors.append(None)
                else:
                    cell = ws.cell(row=r, column=i + 1)
                    fill = cell.fill
                    col = None
                    rgb = fill.fgColor.rgb if fill.fgColor and hasattr(fill.fgColor, 'rgb') else None
                    if rgb and isinstance(rgb, str):
                        col = '#' + rgb[-6:]
                        # 🟢 FIX: IGNORĂ NEGRUL (numai dacă nu vrei să ai negru intenționat ca fundal)
                        if col.lower() == '#000000':
                            col = None
                    row_colors.append(col)
            colors.append(row_colors)
        return colors
    except Exception as e:
        print(f"Error reading colors from Excel file: {e}")
        return []
import os
import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

import os
import time
import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font


def update_excel_from_db():
    from openpyxl.comments import Comment  # import local, ca să fii sigur

    with excel_write_lock:
        try:
            print("\n===== update_excel_from_db =====")
            excel_path = get_active_excel_path()
            sheet_name = get_active_sheet()
            print("[DEBUG] Excel path:", excel_path)
            print("[DEBUG] Sheet name:", sheet_name)
            print("[DEBUG] Există baza DB?", os.path.exists(DB_PATH))
            if not os.path.exists(DB_PATH):
                print("[FATAL] Lipsă DB!")
                return

            # Citește datele din DB
            conn = sqlite3.connect(DB_PATH)
            try:
                df = pd.read_sql(f'SELECT * FROM {sheet_name}', conn)
                print("[DEBUG] Primele rânduri DB:\n", df.head())
            except Exception as e:
                print("[FATAL] Eroare la citire DF:", e)
                return
            finally:
                conn.close()

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Exclude rowid (cheia primară internă) din lista de coloane pentru
            # afișare și la export/export to Excel:
            columns = [col for col in df.columns if not col.endswith('_color') and not col.endswith('_comment') and col.lower() != 'rowid']
            print("[DEBUG] Coloane de bază:", columns)

            header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            alignment = Alignment(horizontal="center", vertical="center")

            for col_num, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment

            # ====================== DATE + CULORI + COMMENT =====================
            for row_num, row_data in enumerate(df[columns].values, 2):
                for col_num, col_name in enumerate(columns, 1):
                    value = row_data[col_num-1]
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    # --- Colorare ---
                    color_col = f"{col_name}_color"
                    if color_col in df.columns:
                        raw_color = df.at[row_num-2, color_col]
                        if (
                            pd.notna(raw_color)
                            and isinstance(raw_color, str)
                            and raw_color.startswith("#")
                            and len(raw_color) == 7
                            and raw_color.upper() != "#FFFFFF"
                        ):
                            cell.fill = PatternFill(start_color=raw_color[1:], end_color=raw_color[1:], fill_type="solid")
                        else:
                            cell.fill = PatternFill(fill_type=None)
                    else:
                        cell.fill = PatternFill(fill_type=None)
                    # --- COMMENT ---
                    comment_col = f"{col_name}_comment"
                    if comment_col in df.columns:
                        comment_txt = df.at[row_num-2, comment_col]
                        if pd.notna(comment_txt) and str(comment_txt).strip():
                            cell.comment = Comment(str(comment_txt), "web") # sau alt autor, după preferință
                        else:
                            cell.comment = None
                    else:
                        cell.comment = None
            # --------------------------------------------------------------------

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        pass
                ws.column_dimensions[column].width = max_length + 2

            tmp_path = excel_path + ".tmp"

            # Test dacă excel_path e blocat/deschis
            if os.path.exists(excel_path):
                try:
                    with open(excel_path, "a"):
                        pass
                except Exception as er:
                    print("[ERROR] Fișierul EXCEL este deschis în altă aplicație! Închide-l și încearcă din nou.")
                    return

            wb.save(tmp_path)
            print("[DEBUG] Salvare temporară OK:", tmp_path)

            # ========  Retry la os.replace cu sleep ========
            for attempt in range(5):
                try:
                    os.replace(tmp_path, excel_path)
                    print(f"[DEBUG] Excel actualizat cu succes! {excel_path}")
                    print("===== FINAL update_excel_from_db =====\n")
                    break
                except PermissionError as e:
                    print(f"[WARNING] Excel lock ({e})! Retry in 1s (attempt {attempt+1}/5)")
                    time.sleep(1)
            else:
                print(f"[FATAL] Nu am putut suprascrie {excel_path} după 5 încercări!")
                # Poți eventual trimite mesaj utilizator dacă vrei
                return

        except Exception as e:
            print(f"[FATAL] Error updating Excel file: {e}")
@app.route('/')
def rel():
    try:
        # --- 1. Setează contextul principal ---
        current_file_id = session.get('file_id', '')
        current_file_name = session.get('file_name', EXCEL_PATH)
        current_sheet = session.get('sheet', SHEET_NAME)
        current_label = 'REL (principal)'

        # --- 2. Creează lista de taburi ---
        tab_list = [{
            'type': 'main',
            'file_id': '',
            'file_name': EXCEL_PATH,
            'sheet': SHEET_NAME,
            'label': 'REL (principal)'
        }]

        filesheets = session.get('filesheets', [])
        for entry in filesheets:
            for sh in entry.get('sheets', []):
                label = f"{sh} ({entry['file_name']})"
                tab_list.append({
                    'type': 'secondary',
                    'file_id': entry['file_id'],
                    'file_name': entry['file_name'],
                    'sheet': sh,
                    'label': label
                })

        # --- 3. Citește datele pentru tabelul activ ---
        if not current_file_id:
            excel_path = EXCEL_PATH
        else:
            excel_path = f"files/{current_file_id}_{current_file_name}"
        sheet_name = current_sheet

        # --- 4. Citește și procesează datele din sheet ---
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=0)
        df.dropna(how='all', inplace=True)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

        columns = [col for col in df.columns if not col.endswith('_color') and not col.endswith('_comment')]
        data = []
        comments = []

        for idx, row in df.iterrows():
         data.append([row[col] for col in columns])  # <-- trebuie să ai și `data`, nu doar `comments`
         comment_row = []
        for col in columns:
         comment_col = f"{col}_comment"
        comment_val = row[comment_col] if comment_col in df.columns else None
        comment_row.append(comment_val)
        comments.append(comment_row)   

        DATE_COLUMNS = [
            "Last calibration date",
            "Calibration Due date",
            "Arrival date of equip",
            "Maintenance planning",
            "Intermediate check planning"
        ]
        DATE_FORMAT = "%d-%b-%y"
        for col in DATE_COLUMNS:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime(DATE_FORMAT)

        df = df.fillna("N/A")
        columns = [col for col in df.columns if not col.endswith('_color')]
        data = df[columns].values.tolist()
        rowids = df.index.tolist()

        # --- ADAUGĂ: REMINDERS (notificări pentru 7/1 zile)
        import datetime
        now = datetime.date.today()
        REMINDER_FIELDS = [
            ("Calibration Due date", "Calibration date"),
            ("Maintenance planning", "Maintenance date"),
        ]
        reminders = []

        for idx, row in df.iterrows():
            equip = row.get("Equipment Identification", "Unknown equipment")
            for field, label in REMINDER_FIELDS:
                raw = row.get(field)
                if not raw or pd.isna(raw):
                    continue
                try:
                    d = pd.to_datetime(raw, errors="coerce")
                    if pd.isnull(d):
                        continue
                    zile = (d.date() - now).days
                    if zile in [7, 1]:
                        msg = f"<b>{equip}</b>: {label}: <b>{d.date()}</b> (<b>{zile} day{'s' if zile == 7 else ''} left</b>)"
                        reminders.append(msg)
                except Exception:
                    pass

        # --- 5. Citește culorile din sheet și ignoră negrul implicit ---
        colors = []
        try:
            wb = load_workbook(excel_path)
            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                row_colors = []
                for cidx in range(1, len(columns) + 1):
                    cell = ws.cell(row=r, column=cidx)
                    fill = cell.fill
                    rgb = fill.fgColor.rgb if fill.fgColor and hasattr(fill.fgColor, 'rgb') else None
                    col = '#' + rgb[-6:] if rgb and isinstance(rgb, str) else None
                    if col and col.lower() == "#000000":
                        col = None  # Fix: NU pune negru implicit
                    row_colors.append(col)
                colors.append(row_colors)
        except Exception as e:
            print("EXCEL COLORS ERROR:", e)
            colors = [["" for _ in columns] for _ in data]

        # --- 6. Render template ---
        return render_template(
            'rel.html',
            columns=columns,
            data=data,
            rowids=rowids,
            colors=colors,
            comments=comments,
            tabs=tab_list,
            current_file_id=current_file_id,
            current_sheet=current_sheet,
            current_file_name=current_file_name,
            reminders=reminders  # <-- ADĂUGAT!
        )
    except Exception as e:
        print(f"Error fetching data for rendering: {e}")
        return f"Error fetching data - {e}", 500
@app.route('/edit_cell', methods=['POST'])
def edit_cell():
    try:
        rowid = request.form['rowid']
        column = request.form['column']
        value = request.form['value']
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(f'UPDATE {SHEET_NAME} SET [{column}] = ? WHERE rowid = ?', (value, rowid))
            conn.commit()
        update_excel_from_db()
        log_change("EDIT", rowid, column, value)
        return jsonify(success=True)
    except Exception as e:
        print(f"Error editing cell: {e}")
        return jsonify(success=False, error=str(e))

@app.route('/add_row', methods=['POST'])
def add_row():
    try:
        user_columns = request.form.getlist('columns[]')
        user_values = request.form.getlist('values[]')

        # Citește toate coloanele actuale din DB (inclusiv _color)
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute(f"PRAGMA table_info({SHEET_NAME})")
            all_cols = [row[1] for row in cur.fetchall()]

        # Pregătește valorile pentru toți parametrii corect: principale + _color
        insert_cols = []
        insert_vals = []
        for c in all_cols:
            insert_cols.append(c)
            if c in user_columns:
                insert_vals.append(user_values[user_columns.index(c)])
            elif c.endswith("_color"):
                insert_vals.append(None)
            else:
                insert_vals.append(None)
        # Fă SQL-ul complet
        placeholders = ','.join(['?'] * len(insert_cols))
        sql = f'INSERT INTO {SHEET_NAME} ({",".join("["+c+"]" for c in insert_cols)}) VALUES ({placeholders})'
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(sql, insert_vals)
            conn.commit()
        update_excel_from_db()
        log_change("ADD", "new", None, str(user_values))
        return jsonify(success=True)
    except Exception as e:
        print(f"Error adding row: {e}")
        return jsonify(success=False, error=str(e))

@app.route('/edit_color', methods=['POST'])
def edit_color():
    try:
        rowid = request.form['rowid']
        column = request.form['column']
        color = request.form['color']
        print("SAVING COLOR:", rowid, column, color)   # <--- AICI
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            if f"{column}_color" not in [col[1] for col in cursor.execute(f"PRAGMA table_info({SHEET_NAME})")]:
                cursor.execute(f"ALTER TABLE {SHEET_NAME} ADD COLUMN [{column}_color] TEXT")
            cursor.execute(f"UPDATE {SHEET_NAME} SET [{column}_color] = ? WHERE rowid = ?", (color, rowid))
            conn.commit()
        update_excel_from_db()
        log_change("COLOR", rowid, column, color)
        return jsonify(success=True)
    except Exception as e:
        print(f"Error editing color: {e}")
        return jsonify(success=False, error=str(e))

@app.route('/download_excel')
def download_excel():
    try:
        update_excel_from_db()
        return send_file(EXCEL_PATH, as_attachment=True)
    except Exception as e:
        print(f"Error downloading Excel: {e}")
        return "Error downloading Excel", 500
@app.route('/add_column', methods=['POST'])
def add_column():
    import re
    column = request.form['column'].strip()
    if not column or not re.match(r'^[\w \-]+$', column):
        return jsonify(success=False, error="Coloana are un nume invalid!")
    try:
        with sqlite3.connect(DB_PATH) as conn:
            # VERIFICĂ EXISTENȚA coloanei!
            cursor = conn.execute(f"PRAGMA table_info({SHEET_NAME})")
            existing_cols = [row[1] for row in cursor.fetchall()]
            if column in existing_cols or f"{column}_color" in existing_cols:
                return jsonify(success=False, error="Coloana există deja!")
            # Adaugă TEXT + _color
            conn.execute(f"ALTER TABLE {SHEET_NAME} ADD COLUMN [{column}] TEXT")
            conn.execute(f"ALTER TABLE {SHEET_NAME} ADD COLUMN [{column}_color] TEXT")
        update_excel_from_db()
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e))
import uuid

#@app.route('/upload_excel', methods=['POST'])
#def upload_excel():
  # file = request.files['file']
 #if file and file.filename.endswith('.xlsx'):
  #      if not os.path.exists('files'):
   #         os.makedirs('files')
    #    fid = uuid.uuid4().hex[:8]
     #   safe_name = secure_filename(file.filename)
      #  filename = f"files/{fid}_{safe_name}"
        #file.save(filename)

        #wb = load_workbook(filename)
        #sheet_list = wb.sheetnames

        # actualizezi sesiunea cu lista de fișiere și sheets
        #filesheets = session.get('filesheets', [])
        #filesheets.append({'file_id': fid, 'file_name': safe_name, 'sheets': sheet_list})
        #session['filesheets'] = filesheets

        #return jsonify(success=True, file_id=fid, sheets=sheet_list, file_name=safe_name)
    #print("Invalid file type uploaded.")
    #return jsonify(success=False), 400
@app.route('/get_history')
def get_history():
    try:
        if os.path.exists(HISTORY_PATH):
            with open(HISTORY_PATH, "r", encoding="utf-8") as f:
                return f.read()
        return "Istoric gol."
    except Exception as e:
        print(f"Error reading history: {e}")
        return "Error fetching history", 500
@app.route('/delete_row', methods=['POST'])
def delete_row():
    try:
        rowid = request.form['rowid']
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(f"DELETE FROM {SHEET_NAME} WHERE rowid=?", (rowid,))
            conn.commit()
        update_excel_from_db()
        log_change("DELETE", rowid)
        return jsonify(success=True)
    except Exception as e:
        print(f"Error deleting row: {e}")
        return jsonify(success=False, error=str(e))
@app.route('/edit_comment', methods=['POST'])
def edit_comment():
    try:
        rowid = request.form['rowid']
        column = request.form['column']
        comment = request.form['comment']
        comment_col = f"{column}_comment"
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            # Adaugă coloana dacă nu există deja:
            existing = [col[1] for col in cursor.execute(f"PRAGMA table_info({SHEET_NAME})")]
            if comment_col not in existing:
                cursor.execute(f"ALTER TABLE {SHEET_NAME} ADD COLUMN [{comment_col}] TEXT")
            cursor.execute(f"UPDATE {SHEET_NAME} SET [{comment_col}] = ? WHERE rowid = ?", (comment, rowid))
            conn.commit()
        # update_excel_from_db() dacă vrei să exporti și comentariile
        return jsonify(success=True)
    except Exception as e:
        print(f"Error editing comment: {e}")
        return jsonify(success=False, error=str(e))
@app.route('/delete_column', methods=['POST'])
def delete_column():
    import re
    column = request.form['column'].strip()
    if not column or not re.match(r'^[\w \-]+$', column):
        return jsonify(success=False, error="Nume de coloană invalid!")
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            # Citește toate coloanele
            cur.execute(f"PRAGMA table_info({SHEET_NAME})")
            cols = [row[1] for row in cur.fetchall()]
            # Refuză ștergerea dacă nu există
            if column not in cols:
                return jsonify(success=False, error="Coloana nu există!")

            # Coloane de păstrat (exclus col și col_color)
            keep_cols = [c for c in cols if c != column and c != f"{column}_color"]

            # Pregătește lista de coloane SQL
            keep_cols_sql = ', '.join(f'[{c}]' for c in keep_cols)

            # 1. Creează tabel nou
            tmp_table = SHEET_NAME + '_tmp'
            cur.execute(f"CREATE TABLE {tmp_table} AS SELECT {keep_cols_sql} FROM {SHEET_NAME}")

            # 2. Șterge tabelul vechi
            cur.execute(f"DROP TABLE {SHEET_NAME}")

            # 3. Redenumește tabelul nou la numele original
            cur.execute(f"ALTER TABLE {tmp_table} RENAME TO {SHEET_NAME}")
            conn.commit()

        update_excel_from_db()
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e))
@app.route('/rename_column', methods=['POST'])
def rename_column():
    import re
    old_col = request.form['old_column'].strip()
    new_col = request.form['new_column'].strip()
    if not old_col or not new_col or not re.match(r'^[\w \-]+$', new_col):
        return jsonify(success=False, error="Invalid column name!")
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            # Verifică dacă NOUL nume există deja
            cur.execute(f"PRAGMA table_info({SHEET_NAME})")
            cols = [row[1] for row in cur.fetchall()]
            if new_col in cols or f"{new_col}_color" in cols:
                return jsonify(success=False, error="A column with this name already exists!")
            # Redenumește coloana principală
            cur.execute(f'ALTER TABLE {SHEET_NAME} RENAME COLUMN [{old_col}] TO [{new_col}]')
            # Dacă există și coloana de culoare, redenumește și pe aceea
            color_src = f"{old_col}_color"
            color_dst = f"{new_col}_color"
            if color_src in cols:
                cur.execute(f'ALTER TABLE {SHEET_NAME} RENAME COLUMN [{color_src}] TO [{color_dst}]')
            conn.commit()
        update_excel_from_db()
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e))
 

from flask import session

@app.route('/set_context', methods=['POST'])
def set_context():
    session['file_id'] = request.form['file_id']
    session['file_name'] = request.form['file_name'] # pt afișare
    session['sheet'] = request.form['sheet']
    return jsonify(success=True)
@app.route('/emails_json')
def emails_json():
    try:
        emails = []
        if os.path.exists('emails.txt'):
            with open('emails.txt', 'r', encoding="utf-8") as f:
                emails = [line.strip() for line in f if "@" in line]
        return jsonify(emails=emails)
    except Exception as e:
        return jsonify(emails=[])

@app.route('/add_email', methods=['POST'])
def add_email():
    email = request.form.get("email", "")
    if '@' not in email or '.' not in email:
        return jsonify(success=False, error="Invalid email!")
    try:
        emails = set()
        if os.path.exists("emails.txt"):
            with open("emails.txt", encoding="utf8") as f:
                emails = set(line.strip() for line in f)
        if email in emails:
            return jsonify(success=False, error="Email already present!")
        with open("emails.txt", "a", encoding="utf8") as f:
            f.write(email.strip() + "\n")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e))

@app.route('/delete_email', methods=['POST'])
def delete_email():
    email = request.form.get("email", "")
    try:
        if not os.path.exists("emails.txt"):
            return jsonify(success=False, error="No file found!")
        with open("emails.txt", encoding="utf8") as f:
            lines = f.readlines()
        with open("emails.txt", "w", encoding="utf8") as f:
            for line in lines:
                if line.strip() != email:
                    f.write(line)
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e))
    
if __name__ == '__main__':
    try:
        if not os.path.exists(DB_PATH):
            init_db()
        app.run(host='0.0.0.0', port=5000, debug=True)
    except Exception as e:
        print(f"Error starting server: {e}")
        from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

def generate_formatted_excel():
    try:
        excel_path = get_active_excel_path()
        sheet_name = get_active_sheet()
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql(f'SELECT * FROM {SHEET_NAME}', conn)
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        alignment = Alignment(horizontal="center", vertical="center")

        for col_num, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment

        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(excel_path)
    except Exception as e:
        print(f"Error updating Excel file: {e}")
